"""XLSX loader — lazy ``openpyxl`` import.

Each worksheet becomes text lines (one per row, cells joined with ``" | "``).
Raises a clear :class:`~xyberos.exceptions.provider.ProviderError` when
``openpyxl`` is not installed (``pip install xyberos[documents]``).
"""

from __future__ import annotations

from pathlib import Path

from xyberos.exceptions.provider import ProviderError

from .base import Document, chunk_documents


def _read_xlsx(path: Path) -> tuple[str, list[str], int]:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise ProviderError(
            "the 'openpyxl' package is required to load XLSX files; "
            "install it with 'pip install xyberos[documents]'"
        ) from exc

    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        sheets = workbook.sheetnames
        lines: list[str] = []
        row_count = 0
        for sheet in workbook.worksheets:
            lines.append(f"# {sheet.title}")
            for row in sheet.iter_rows(values_only=True):
                row_count += 1
                cells = ["" if cell is None else str(cell) for cell in row]
                lines.append(" | ".join(cells))
        return "\n".join(lines), sheets, row_count
    finally:
        workbook.close()


class XlsxLoader:
    """Loads an Excel ``.xlsx`` file as text (one line per row per sheet)."""

    def __init__(self, chunk_size: int | None = None) -> None:
        self._chunk_size = chunk_size

    def load(self, path: str) -> list[Document]:
        path_obj = Path(path)
        if not path_obj.is_file():
            raise FileNotFoundError(path)
        text, sheets, row_count = _read_xlsx(path_obj)
        document = Document(
            source=str(path_obj),
            text=text,
            metadata={"sheets": sheets, "row_count": row_count, "extension": ".xlsx"},
        )
        return chunk_documents([document], self._chunk_size)
